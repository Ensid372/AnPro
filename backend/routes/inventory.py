from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, InventoryRecord, InventoryItem, Component, ReadyMedicine, User
from datetime import datetime, date
import io

inventory_bp = Blueprint('inventory', __name__)

@inventory_bp.route('/', methods=['GET'])
@jwt_required()
def list_inventories():
    records = InventoryRecord.query.order_by(InventoryRecord.created_at.desc()).all()
    return jsonify([r.to_dict() for r in records])

@inventory_bp.route('/<int:rid>', methods=['GET'])
@jwt_required()
def get_inventory(rid):
    record = InventoryRecord.query.get_or_404(rid)
    data = record.to_dict()
    data['items'] = [i.to_dict() for i in record.items]
    return jsonify(data)

@inventory_bp.route('/start', methods=['POST'])
@jwt_required()
def start_inventory():
    user_id = get_jwt_identity()
    today = date.today()

    record = InventoryRecord(
        conducted_by=int(user_id),
        status='draft'
    )
    db.session.add(record)
    db.session.flush()

    for comp in Component.query.all():
        is_expired = bool(comp.expiry_date and comp.expiry_date < today)
        is_below = comp.quantity < comp.critical_norm
        item = InventoryItem(
            record_id=record.id,
            item_type='component',
            item_id=comp.id,
            item_name=comp.name,
            expected_quantity=comp.quantity,
            actual_quantity=comp.quantity,
            critical_norm=comp.critical_norm,
            expiry_date=comp.expiry_date,
            is_expired=is_expired,
            is_below_critical=is_below,
            discrepancy=0.0
        )
        db.session.add(item)

    for med in ReadyMedicine.query.all():
        is_expired = bool(med.expiry_date and med.expiry_date < today)
        is_below = med.quantity < med.critical_norm
        item = InventoryItem(
            record_id=record.id,
            item_type='ready_medicine',
            item_id=med.id,
            item_name=med.name,
            expected_quantity=med.quantity,
            actual_quantity=med.quantity,
            critical_norm=med.critical_norm,
            expiry_date=med.expiry_date,
            is_expired=is_expired,
            is_below_critical=is_below,
            discrepancy=0.0
        )
        db.session.add(item)

    db.session.commit()
    data = record.to_dict()
    data['items'] = [i.to_dict() for i in record.items]
    return jsonify(data), 201

@inventory_bp.route('/<int:rid>/item/<int:iid>', methods=['PUT'])
@jwt_required()
def update_inventory_item(rid, iid):
    item = InventoryItem.query.get_or_404(iid)
    data = request.get_json()

    if 'actual_quantity' in data:
        item.actual_quantity = float(data['actual_quantity'])
        item.discrepancy = item.actual_quantity - (item.expected_quantity or 0)
        item.is_below_critical = item.actual_quantity < (item.critical_norm or 0)

    db.session.commit()
    return jsonify(item.to_dict())

@inventory_bp.route('/<int:rid>/complete', methods=['POST'])
@jwt_required()
def complete_inventory(rid):
    record = InventoryRecord.query.get_or_404(rid)
    data = request.get_json() or {}
    record.status = 'completed'
    record.completed_at = datetime.utcnow()
    record.notes = data.get('notes', '')

    apply_changes = data.get('apply_changes', True)

    if apply_changes:
        for item in record.items:
            if item.item_type == 'component':
                comp = Component.query.get(item.item_id)
                if comp:
                    comp.quantity = item.actual_quantity
            elif item.item_type == 'ready_medicine':
                med = ReadyMedicine.query.get(item.item_id)
                if med:
                    med.quantity = item.actual_quantity

    db.session.commit()
    return jsonify(record.to_dict())

@inventory_bp.route('/report', methods=['GET'])
@jwt_required()
def inventory_report():
    today = date.today()

    components = Component.query.all()
    ready_medicines = ReadyMedicine.query.all()

    below_critical = []
    expired = []

    for comp in components:
        if comp.quantity < comp.critical_norm:
            below_critical.append({
                'type': 'component',
                'name': comp.name,
                'quantity': comp.quantity,
                'critical_norm': comp.critical_norm,
                'unit': comp.unit
            })
        if comp.expiry_date and comp.expiry_date < today:
            expired.append({
                'type': 'component',
                'name': comp.name,
                'quantity': comp.quantity,
                'expiry_date': comp.expiry_date.isoformat(),
                'unit': comp.unit
            })

    for med in ready_medicines:
        if med.quantity < med.critical_norm:
            below_critical.append({
                'type': 'ready_medicine',
                'name': med.name,
                'quantity': med.quantity,
                'critical_norm': med.critical_norm,
                'unit': med.unit
            })
        if med.expiry_date and med.expiry_date < today:
            expired.append({
                'type': 'ready_medicine',
                'name': med.name,
                'quantity': med.quantity,
                'expiry_date': med.expiry_date.isoformat(),
                'unit': med.unit
            })

    return jsonify({
        'generated_at': datetime.utcnow().isoformat(),
        'below_critical': below_critical,
        'expired': expired,
        'total_below_critical': len(below_critical),
        'total_expired': len(expired)
    })

@inventory_bp.route('/statistics', methods=['GET'])
@jwt_required()
def statistics():
    from models import Order
    from sqlalchemy import func

    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    query = Order.query.filter(Order.status == 'completed')

    if date_from:
        query = query.filter(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(Order.created_at <= datetime.fromisoformat(date_to))

    orders = query.all()

    medicine_stats = {}
    for order in orders:
        name = order.medicine_name
        medicine_stats[name] = medicine_stats.get(name, 0) + 1

    stats_list = [
        {'medicine_name': k, 'orders_count': v}
        for k, v in sorted(medicine_stats.items(), key=lambda x: -x[1])
    ]

    return jsonify({
        'period': {'from': date_from, 'to': date_to},
        'total_orders': len(orders),
        'medicine_stats': stats_list
    })

@inventory_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl не установлен'}), 500

    today = date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Инвентаризация'

    ws.merge_cells('A1:G1')
    ws['A1'] = f'ИНВЕНТАРИЗАЦИОННАЯ ВЕДОМОСТЬ — {today.strftime("%d.%m.%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Наименование', 'Тип', 'Количество', 'Ед.изм.', 'Критич. норма', 'Срок годности', 'Статус']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='DCE6F1')

    row = 4
    for comp in Component.query.order_by(Component.name).all():
        is_exp = comp.expiry_date and comp.expiry_date < today
        status = []
        if comp.quantity < comp.critical_norm:
            status.append('Ниже нормы')
        if is_exp:
            status.append('Просрочен')
        ws.cell(row=row, column=1, value=comp.name)
        ws.cell(row=row, column=2, value='Компонент')
        ws.cell(row=row, column=3, value=comp.quantity)
        ws.cell(row=row, column=4, value=comp.unit)
        ws.cell(row=row, column=5, value=comp.critical_norm)
        ws.cell(row=row, column=6, value=comp.expiry_date.strftime('%d.%m.%Y') if comp.expiry_date else '')
        ws.cell(row=row, column=7, value=', '.join(status) if status else 'OK')
        row += 1

    for med in ReadyMedicine.query.order_by(ReadyMedicine.name).all():
        is_exp = med.expiry_date and med.expiry_date < today
        status = []
        if med.quantity < med.critical_norm:
            status.append('Ниже нормы')
        if is_exp:
            status.append('Просрочен')
        ws.cell(row=row, column=1, value=med.name)
        ws.cell(row=row, column=2, value='Готовое ЛС')
        ws.cell(row=row, column=3, value=med.quantity)
        ws.cell(row=row, column=4, value=med.unit)
        ws.cell(row=row, column=5, value=med.critical_norm)
        ws.cell(row=row, column=6, value=med.expiry_date.strftime('%d.%m.%Y') if med.expiry_date else '')
        ws.cell(row=row, column=7, value=', '.join(status) if status else 'OK')
        row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'inventory_{today.isoformat()}.xlsx')
